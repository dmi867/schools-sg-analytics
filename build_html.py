#!/usr/bin/env python3
"""Generate index.html from Schools xlsx data."""
import json
import math
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent

ADVANCE_PCTS = {30.0, 49.0}  # типовые проценты аванса, а не расчётный факт оплаты

# Ручные исправления опечаток источника (Simple List расходится с более свежими данными).
NAME_OVERRIDES = {
    "1000001282.1000001075": "МБОУ СОШ № 28 ГОЩ",  # источник даёт МАОУ, актуальный тип — МБОУ
}


def short(n, keep_prefix=False):
    n = n or ""
    if not keep_prefix:
        n = n.replace("МБОУ ", "").replace("МАОУ ", "").replace("МОУ ", "")
        n = n.lstrip("-– ").strip()
    return (n[:42] + "…") if len(n) > 42 else n


def short_contractor(c):
    if not c:
        return None
    c = re.sub(
        r'(ООО|ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ|ЗАО|АО|Закрытое акционерное общество|Акционерное общество)\s*',
        "", str(c), flags=re.I,
    ).strip()
    c = c.strip('"\' ')
    return c.title() if c.isupper() else c


def parse_date(v):
    if isinstance(v, datetime):
        return v
    if not v:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(v).strip(), fmt)
        except ValueError:
            pass
    return None


def parse_amt(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace("\xa0", "").replace(" ", "").replace(",", "."))
    except ValueError:
        return 0.0


def two_tailed_p(r, n):
    if r is None or n < 3 or abs(r) >= 1:
        return None
    t = r * math.sqrt(n - 2) / math.sqrt(1 - r**2)
    return round(2 * (1 - 0.5 * (1 + math.erf(abs(t) / math.sqrt(2)))), 3)


def median(xs):
    n = len(xs)
    if n == 0:
        return 0
    s = sorted(xs)
    mid = n // 2
    return (s[mid - 1] + s[mid]) / 2 if n % 2 == 0 else s[mid]


def corr(xs, ys):
    n = len(xs)
    if n < 3:
        return None
    mx, my = sum(xs) / n, sum(ys) / n
    num = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    dx = math.sqrt(sum((x - mx) ** 2 for x in xs))
    dy = math.sqrt(sum((y - my) ** 2 for y in ys))
    if dx == 0 or dy == 0:
        return None
    return num / (dx * dy)


def linreg(xs, ys):
    n = len(xs)
    mx, my = sum(xs) / n, sum(ys) / n
    num = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    den = sum((x - mx) ** 2 for x in xs)
    b = num / den
    a = my - b * mx
    pred = [a + b * x for x in xs]
    ss_res = sum((y - p) ** 2 for y, p in zip(ys, pred))
    ss_tot = sum((y - my) ** 2 for y in ys)
    return a, b, 1 - ss_res / ss_tot if ss_tot else 0


def fmt_date(d):
    return d.strftime("%d.%m.%y") if d else None


def kt_bounds(items):
    if not items:
        return {}
    return {
        "plan_start": min((x["ps"] for x in items if x["ps"]), default=None),
        "plan_end": max((x["pe"] for x in items if x["pe"]), default=None),
        "fact_start": min((x["fs"] for x in items if x["fs"]), default=None),
        "fact_end": max((x["fe"] for x in items if x["fe"]), default=None),
    }


def load_ksg():
    path = ROOT / "1708_КСГ+Экспертиза.xlsx"
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    col = {h: i for i, h in enumerate(headers)}
    kt = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        uin = row[col["УИН"]]
        if not uin:
            continue
        proc = row[col["Процедура КСГ"]]
        kt.setdefault(uin, {}).setdefault(proc, []).append(
            {
                "ps": parse_date(row[col["Дата начала план"]]),
                "pe": parse_date(row[col["Дата окончания план"]]),
                "fs": parse_date(row[col["Дата начала факт"]]),
                "fe": parse_date(row[col["Дата окончания факт"]]),
            }
        )
    wb.close()
    return kt


def load_simple_list():
    wb = openpyxl.load_workbook(ROOT / "2408_Акцент_Simple List.xlsx", data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}
    sl = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        uin = row[col["Код УИН"]]
        if not uin:
            continue
        yr = row[col["Год финансирования"]]
        if uin in sl and yr and str(yr).isdigit() and int(yr) <= int(sl[uin].get("yr") or 0):
            continue
        pay_pct = None
        v = row[col["Процент выплат"]]
        if v is not None:
            try:
                pay_pct = round(float(str(v).replace(",", ".")), 1)
            except ValueError:
                pass
        contractor = row[col["Наименование подрядчика"]] or (sl.get(uin) or {}).get("contractor")
        sl[uin] = {
            "yr": yr,
            "name": row[col["Название объекта"]],
            "pay_pct": pay_pct,
            "contractor": contractor,
            "exp_in": parse_date(row[col["Дата подачи заявления (захода) на экспертизу"]]),
            "exp_start": parse_date(row[col["Дата начала экспертизы"]]),
            "exp_done": parse_date(row[col["Дата получения заключения (завершения ) экспертизы"]]),
            "ctr_plan": parse_date(row[col["Заключение контракта начало план КСГ"]]),
            "ctr_fact": parse_date(row[col["Заключение контракта начало факт КСГ"]]),
        }
    return sl


def load_finance2026():
    path = ROOT / "finance2026.json"
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def build_flags(last, info, exp, smr, rs, ctr):
    flags = []
    sg_plan_gap = round(last["fact"] - last["plan"], 1) if last["plan"] is not None else None
    pay_pct = info.get("pay_pct")
    sg_pay_gap = round(last["fact"] - pay_pct, 1) if pay_pct is not None else None

    if sg_pay_gap is not None and sg_pay_gap > 55:
        flags.append("готовность выше выплат")
    if sg_plan_gap is not None and sg_plan_gap < -5:
        flags.append("отстаёт от плана")
    if smr.get("fact_start") and exp.get("plan_end") and smr["fact_start"] < exp["plan_end"]:
        flags.append("стройка до экспертизы")
    if info.get("exp_done") and exp.get("plan_end") and info["exp_done"] < exp["plan_end"]:
        flags.append("экспертиза не бьётся с КСГ")
    if smr.get("fact_start") and ctr.get("fact_start") and smr["fact_start"] < ctr["fact_start"]:
        flags.append("стройка до контракта")
    if rs.get("plan_end") and rs.get("fact_end") and rs["fact_end"] > rs["plan_end"]:
        flags.append("РС опоздал")
    elif not rs and smr.get("fact_start") and not info.get("exp_done"):
        flags.append("нет заключения экспертизы")

    return flags, sg_plan_gap, sg_pay_gap


def load_data():
    sl = load_simple_list()
    ksg = load_ksg()
    fin2026 = load_finance2026()

    cross, per, traj, kt_rows, kt_dates = [], [], {}, [], {}
    budget_alert = []
    STAGE_EDGES = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100]
    stage_gaps = {e: {"plan": [], "pay": []} for e in STAGE_EDGES}

    for f in sorted(ROOT.glob("*.xlsx")):
        if (
            "_платежи" in f.name
            or "Акцент" in f.name
            or "Simple" in f.name
            or "КСГ" in f.name
        ):
            continue
        uin = f.stem
        info = sl.get(uin, {})
        name = NAME_OVERRIDES.get(uin, info.get("name", uin))
        wb = openpyxl.load_workbook(f, data_only=True)
        series = []
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            d = parse_date(row[0])
            if not d:
                continue
            try:
                plan = float(row[1]) if row[1] is not None else None
                fact = float(row[2])
            except (TypeError, ValueError):
                continue
            series.append({"d": d, "plan": plan, "fact": fact})
        series.sort(key=lambda x: x["d"])
        if not series:
            continue

        pays = []
        pf = ROOT / f"{uin}_платежи.xlsx"
        if pf.exists():
            wb2 = openpyxl.load_workbook(pf, data_only=True)
            hdr = [c.value for c in next(wb2.active.iter_rows(min_row=1, max_row=1))]
            if hdr and hdr[0] == "Тип платежа":
                for row in wb2.active.iter_rows(min_row=2, values_only=True):
                    d = parse_date(row[4])
                    amt = parse_amt(row[5])
                    if d and amt:
                        pays.append({"d": d, "amt": amt, "advance": row[0] == "Предоплата"})
        pays.sort(key=lambda x: x["d"])
        total = sum(p["amt"] for p in pays)
        last = series[-1]
        pay_pct = info.get("pay_pct")

        pay_scale = pay_pct / total if (pay_pct is not None and total > 0) else None

        cum = cum_adv = cum_reg = 0
        pi = 0
        pts = []
        for pt in series:
            while pi < len(pays) and pays[pi]["d"] <= pt["d"]:
                amt = pays[pi]["amt"]
                cum += amt
                if pays[pi]["advance"]:
                    cum_adv += amt
                else:
                    cum_reg += amt
                pi += 1
            pts.append(
                {
                    "d": pt["d"].strftime("%d.%m.%y"),
                    "sg": round(pt["fact"], 1),
                    "plan": round(pt["plan"], 1) if pt["plan"] is not None else None,
                    "pay": round(cum / 1e6, 1),
                    "payPct": round(cum * pay_scale, 1) if pay_scale is not None else None,
                    "advPct": round(cum_adv * pay_scale, 1) if pay_scale is not None else None,
                    "regPct": round(cum_reg * pay_scale, 1) if pay_scale is not None else None,
                }
            )
        for edge in STAGE_EDGES:
            hit = next((p for p in pts if p["sg"] >= edge), None)
            if hit is None:
                continue
            if hit["plan"] is not None:
                stage_gaps[edge]["plan"].append(hit["sg"] - hit["plan"])
            if hit["payPct"] is not None:
                stage_gaps[edge]["pay"].append(hit["sg"] - hit["payPct"])

        if len(pts) > 36:
            step = math.ceil(len(pts) / 36)
            pts = pts[::step][:-1] + [pts[-1]]
        traj[uin] = pts

        pays_t = [p["pay"] for p in pts]
        vary = max(pays_t) - min(pays_t) >= 0.01
        r = None
        if vary and len(pts) >= 5:
            r = corr([p["sg"] for p in pts], pays_t)

        uin_kt = ksg.get(uin, {})
        exp = kt_bounds(uin_kt.get("Экспертиза", []))
        smr = kt_bounds(uin_kt.get("СМР", []))
        ctr = kt_bounds(uin_kt.get("Заключение контракта 1", []))
        rs = kt_bounds(uin_kt.get("Получение РС", []))
        flags, sg_plan_gap, sg_pay_gap = build_flags(last, info, exp, smr, rs, ctr)

        fin = fin2026.get(uin)
        if fin and fin.get("osv2026") == 0:
            flags.append("не осваивает бюджет 2026")

        exp_ref = info.get("exp_done") or exp.get("plan_end")
        exp_marker = None
        if exp_ref and pts:
            exp_marker = min(
                pts, key=lambda p: abs((datetime.strptime(p["d"], "%d.%m.%y") - exp_ref).days)
            )["d"]

        kt_dates[uin] = {
            "exp_plan": fmt_date(exp.get("plan_end")),
            "exp_fact": fmt_date(exp.get("fact_end")),
            "exp_sl": fmt_date(info.get("exp_done")),
            "exp_marker": exp_marker,
            "smr_start": fmt_date(smr.get("fact_start")),
            "ctr_fact": fmt_date(ctr.get("fact_start") or info.get("ctr_fact")),
            "rs_plan": fmt_date(rs.get("plan_end")),
            "rs_fact": fmt_date(rs.get("fact_end")),
        }

        contractor = short_contractor(info.get("contractor"))

        kt_rows.append(
            {
                "uin": uin,
                "name": short(name),
                "full": name,
                "sg": round(last["fact"], 1),
                "plan": round(last["plan"], 1) if last["plan"] is not None else None,
                "sg_plan_gap": sg_plan_gap,
                "pct": pay_pct,
                "sg_pay_gap": sg_pay_gap,
                "flags": flags,
                "flag_n": len(flags),
                "contractor": contractor,
            }
        )

        cross.append(
            {
                "uin": uin,
                "name": short(name),
                "full": name,
                "sg": round(last["fact"], 1),
                "pct": pay_pct,
                "advance": pay_pct in ADVANCE_PCTS,
                "pay": round(total / 1e6, 1),
                "gap": sg_pay_gap,
                "contractor": contractor,
            }
        )
        per.append(
            {
                "uin": uin,
                "name": short(name),
                "sg": round(last["fact"], 1),
                "plan": round(last["plan"], 1) if last["plan"] is not None else None,
                "sg_plan_gap": sg_plan_gap,
                "pct": pay_pct,
                "pay": round(total / 1e6, 1),
                "r": round(r, 3) if r is not None else None,
                "vary": vary,
                "n": len(series),
                "flags": flags,
            }
        )

    name_counts = Counter(c["name"] for c in cross)
    dup_uins = {c["uin"] for c in cross if name_counts[c["name"]] > 1}
    if dup_uins:
        uin_full = {c["uin"]: c["full"] for c in cross}
        for row in (*kt_rows, *cross, *per):
            if row["uin"] in dup_uins:
                row["name"] = short(uin_full[row["uin"]], keep_prefix=True)

    for k in kt_rows:
        fin = fin2026.get(k["uin"])
        if fin and fin.get("osv2026") == 0:
            budget_alert.append(
                {
                    "uin": k["uin"],
                    "name": k["name"],
                    "full": k["full"],
                    "sg": k["sg"],
                    "plan2026": round(fin["plan2026_rub"] / 1e6, 1),
                }
            )
    budget_alert.sort(key=lambda x: -x["plan2026"])

    by_contractor = {}
    for k in kt_rows:
        c = k["contractor"] or "Не указан"
        e = by_contractor.setdefault(c, {"contractor": c, "objects": [], "n_no_budget2026": 0, "n_flags": 0})
        no_bud = "не осваивает бюджет 2026" in k["flags"]
        e["objects"].append({"name": k["name"], "full": k["full"], "sg": k["sg"], "no_budget2026": no_bud})
        if no_bud:
            e["n_no_budget2026"] += 1
        if k["flags"]:
            e["n_flags"] += 1
    contractors = sorted(by_contractor.values(), key=lambda x: (-len(x["objects"]), x["contractor"]))

    stage_analysis = []
    for edge in STAGE_EDGES:
        plan_vals = stage_gaps[edge]["plan"]
        pay_vals = stage_gaps[edge]["pay"]
        if not plan_vals and not pay_vals:
            continue
        stage_analysis.append(
            {
                "stage": edge,
                "n": max(len(plan_vals), len(pay_vals)),
                "plan_gap": round(median(plan_vals), 1) if plan_vals else None,
                "pay_gap": round(median(pay_vals), 1) if pay_vals else None,
            }
        )

    valid = [s for s in cross if s["pct"] is not None and s["pay"] > 0]
    facts = [s["sg"] for s in valid]
    pcts = [s["pct"] for s in valid]
    a, b, r2 = linreg(pcts, facts)
    varying = [p for p in per if p["vary"] and p["r"] is not None]
    varying.sort(key=lambda x: -(x["r"] or -1))
    rs_corr = [p["r"] for p in varying]

    gaps = [s["gap"] for s in cross if s["gap"] is not None]
    kt_sorted = sorted(kt_rows, key=lambda x: (-x["flag_n"], -(x["sg_pay_gap"] or 0)))

    no_adv = [s for s in valid if not s["advance"]]
    facts_na = [s["sg"] for s in no_adv]
    pcts_na = [s["pct"] for s in no_adv]
    pearson_na = corr(facts_na, pcts_na) if len(no_adv) >= 3 else None

    return {
        "stats": {
            "n": len(valid),
            "median_r": round(median(rs_corr), 3) if rs_corr else 0,
            "n_varying": len(varying),
            "n_sg_ahead": sum(1 for g in gaps if g > 55),
            "median_gap": round(median(gaps), 1) if gaps else 0,
            "n_sg_behind_plan": sum(
                1 for p in per if p.get("sg_plan_gap") is not None and p["sg_plan_gap"] < -5
            ),
            "n_smr_before_exp": sum(1 for p in per if "стройка до экспертизы" in p.get("flags", [])),
            "n_kt_issues": sum(1 for p in per if any(f in p.get("flags", []) for f in ("стройка до экспертизы", "экспертиза не бьётся с КСГ", "РС опоздал", "нет заключения экспертизы"))),
            "n_no_advance": len(no_adv),
            "pearson_no_advance": round(pearson_na, 3) if pearson_na is not None else None,
            "pearson_no_advance_p": two_tailed_p(pearson_na, len(no_adv)),
            "n_no_budget2026": len(budget_alert),
        },
        "budget_alert": budget_alert,
        "contractors": contractors,
        "stage_analysis": stage_analysis,
        "kt_attention": kt_sorted[:10],
        "reg": {"a": round(a, 1), "b": round(b, 4)},
        "cross": cross,
        "per": sorted(per, key=lambda x: -(x["r"] if x["r"] is not None else -1)),
        "traj": traj,
        "kt_dates": kt_dates,
        "defaultUin": varying[0]["uin"] if varying else per[0]["uin"],
    }


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>СГ и выплаты — 47 школ</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.2.0/dist/chartjs-plugin-datalabels.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-annotation@3.0.1/dist/chartjs-plugin-annotation.min.js"></script>
<style>
  :root { --bg:#f4f3ef; --surface:#fff; --text:#1c1c1c; --muted:#555; --faint:#888;
    --line:#ddd9d0; --accent:#1a4d7a; --green:#2d6a4f;
    --warn-bg:#faf3e6; --warn-border:#c4922a; --info-bg:#edf3f9; --ok-bg:#eaf4ee; }
  * { box-sizing:border-box }
  body { margin:0; font:15px/1.5 "Segoe UI",system-ui,sans-serif; background:var(--bg); color:var(--text) }
  .wrap { max-width:880px; margin:0 auto; padding:24px 18px 56px }
  h1 { font-size:1.5rem; font-weight:650; margin:0 0 4px }
  .sub { color:var(--muted); margin:0 0 20px; font-size:.92rem }
  ul.brief { margin:0; padding-left:1.2rem; color:var(--muted) }
  ul.brief li { margin:6px 0 }
  .kpis { display:grid; grid-template-columns:repeat(3,1fr); gap:10px; margin:16px 0 }
  .kpi { background:var(--surface); border:1px solid var(--line); border-radius:6px; padding:12px 14px }
  .kpi .n { font-size:1.35rem; font-weight:700; color:var(--accent); line-height:1.2 }
  .kpi .l { font-size:.75rem; color:var(--muted); margin-top:4px }
  .chart { position:relative; height:280px; margin-top:10px }
  .chart.tall { height:340px }
  .note { font-size:.82rem; color:var(--faint); margin:6px 0 0 }
  .box { background:var(--surface); border:1px solid var(--line); border-radius:6px; padding:14px 16px; margin:12px 0 }
  .mini table { width:100%; font-size:.84rem; border-collapse:collapse }
  .mini th,.mini td { padding:6px 8px; border-bottom:1px solid var(--line); text-align:left }
  .mini th { color:var(--muted); font-weight:600 }
  .mini td.r,.mini th.r { text-align:right; font-variant-numeric:tabular-nums }
  details { background:var(--surface); border:1px solid var(--line); border-radius:6px; margin:10px 0; overflow:hidden }
  details > summary { cursor:pointer; padding:14px 16px; font-weight:600; list-style:none; user-select:none }
  details > summary::-webkit-details-marker { display:none }
  details > summary::after { content:'+'; float:right; color:var(--faint); font-weight:400 }
  details[open] > summary::after { content:'−' }
  details > summary:hover { background:#faf9f6 }
  .detail-body { padding:0 16px 16px; border-top:1px solid var(--line) }
  select { font:inherit; padding:6px 10px; border:1px solid var(--line); border-radius:4px; background:#fff; max-width:100% }
  .row { display:flex; gap:8px; align-items:center; flex-wrap:wrap; margin:10px 0 }
  .tag { font-size:.75rem; padding:3px 9px; background:var(--info-bg); border:1px solid var(--line); border-radius:99px; color:var(--accent) }
  table.full { width:100%; border-collapse:collapse; font-size:.84rem }
  table.full th,table.full td { padding:7px 9px; border-bottom:1px solid var(--line) }
  table.full th { background:#eeebe4; position:sticky; top:0; text-align:left }
  table.full td.r,table.full th.r { text-align:right; font-variant-numeric:tabular-nums }
  .tbl-wrap { max-height:420px; overflow:auto; border:1px solid var(--line); border-radius:4px; margin-top:10px }
  .flag { font-size:.72rem; padding:2px 7px; margin:1px 2px 1px 0; display:inline-block; background:#faf3e6; border:1px solid #e0c88a; border-radius:4px; color:#7a5a12 }
  .flag.warn { background:#fdecea; border-color:#e8a8a0; color:#8a2a22 }
  .kpis.four { grid-template-columns:repeat(4,1fr) }
  @media(max-width:900px) { .kpis.four { grid-template-columns:repeat(2,1fr) } }
  @media(max-width:700px) { .kpis,.kpis.four { grid-template-columns:1fr } }
</style>
</head>
<body>
<div class="wrap">
  <h1>СГ и выплаты</h1>
  <p class="sub">47 школ, август 2026</p>

  <ul class="brief">
    <li>У <span id="b1"></span> школ готовность ниже плана больше чем на 5 п.п.</li>
    <li>У <span id="b2"></span> школ готовность заметно выше того, что уже заплатили.</li>
    <li>У <span id="b3"></span> школ стройка началась раньше, чем в КСГ стоит экспертиза.</li>
    <li><strong>У <span id="b4"></span> школ в 2026 году не потрачено ни рубля из утверждённого бюджета</strong> — см. блок ниже.</li>
    <li>Про РС: в КСГ по школам такой точки нет, смотрим контракт и экспертизу.</li>
  </ul>

  <div class="kpis four">
    <div class="kpi"><div class="n" id="k1"></div><div class="l">готовность выше выплат</div></div>
    <div class="kpi"><div class="n" id="k2"></div><div class="l">обычный разрыв, п.п.</div></div>
    <div class="kpi"><div class="n" id="k4"></div><div class="l">факт ниже плана</div></div>
    <div class="kpi"><div class="n" id="k5"></div><div class="l">стройка до экспертизы</div></div>
  </div>

  <div class="box" style="background:#fdecea;border-color:#e8a8a0">
    <strong style="display:block;margin-bottom:6px;color:#8a2a22">⚠ <span id="k6"></span> школ не осваивают бюджет 2026 года — деньги утверждены, но не выплачены</strong>
    <p class="note" style="margin:0 0 8px">Это не аванс и не обычная задержка — по этим объектам с начала 2026 года кассовых выплат не было вообще (0%), при этом строительство почти завершено (готовность 80–100%). Деньги на этот год утверждены, просто не идут.</p>
    <table>
      <thead><tr><th>Школа</th><th class="r">СГ</th><th class="r">План 2026, млн ₽</th></tr></thead>
      <tbody id="budgetAlert"></tbody>
    </table>
  </div>

  <div class="box">
    <strong style="display:block;margin-bottom:8px">Главный вывод: разрыв растёт по ходу стройки, а не постоянный</strong>
    <p class="note" style="margin:0 0 8px">По всем 48 школам смотрим не на календарные даты, а на этап готовности (10%, 20%, ... 100%) — и в момент, когда объект впервые доходит до этого этапа, смотрим, на сколько план и деньги от него отстают. Так можно сравнивать разные школы на одной шкале и увидеть, где именно теряются деньги и время.</p>
    <div class="chart tall"><canvas id="cStage"></canvas></div>
    <p class="note" id="stageNote" style="margin-top:8px"></p>
  </div>

  <details id="secMethod">
    <summary>Методика управления финансированием портфеля</summary>
    <div class="detail-body">
      <p class="note" style="margin-top:14px">Ниже — не мониторинг стройготовности, а взгляд с точки зрения денег: когда давать финансирование, когда придержать и на что смотреть в первую очередь. Построено на графике выше и на разрезах по школам/подрядчикам ниже.</p>

      <strong style="display:block;margin-top:14px">1. Эталонная кривая вместо разовой цифры</strong>
      <p class="note">Разрыв «СГ минус выплаты» — это не ошибка и не всегда плохо: он закономерно растёт по ходу стройки почти у всех объектов (аванс в начале — это нормально, оплата по актам после экспертизы отстаёт — тоже ожидаемо). Ориентир — не «оплачено меньше 100%», а «отклоняется ли конкретный объект от того, что типично для его этапа готовности».</p>

      <strong style="display:block;margin-top:14px">2. Правило красной зоны</strong>
      <p class="note">На этапе ~90% готовности типичный разрыв по порталу — около <span id="mth90"></span> п.п. Если у объекта на том же этапе разрыв заметно больше — это уже не типовой лаг платёжного цикла, а сигнал разобраться отдельно: не подаются акты, спор с подрядчиком, затор у заказчика.</p>

      <strong style="display:block;margin-top:14px">3. Нулевое освоение — отдельный, более жёсткий триггер</strong>
      <p class="note"><span id="mth0"></span> школ показывают 0% кассового исполнения бюджета текущего года при уже утверждённом финансировании — это не про этап готовности вообще, деньги просто не двигаются. Такие объекты не ждут конца года — они эскалируются сразу, отдельным списком (см. блок выше).</p>

      <strong style="display:block;margin-top:14px">4. Аванс — это не «дать ещё денег»</strong>
      <p class="note">Если по объекту идут только авансовые платежи и ни одного основного (по актам) — значит, вопрос не в объёме финансирования, а в том, что документы о выполненных работах не доходят до оплаты. Дальнейшее решение — не наращивать аванс, а разбираться, почему не закрываются акты. Ориентир по порталу: аванс оправдан на ранней стадии (примерно до 30–40% готовности), дальше приоритет должен смещаться на основные платежи.</p>

      <strong style="display:block;margin-top:14px">5. Смотреть на подрядчика, а не только на объект</strong>
      <p class="note" id="mthContr">—</p>

      <strong style="display:block;margin-top:14px">Что это даёт на практике</strong>
      <ul class="brief">
        <li>Перед очередным траншем — сверять этап готовности объекта с эталонной кривой, а не только со сроком по контракту.</li>
        <li>Держать список нулевого освоения как еженедельный контроль-лист для эскалации, а не ждать годового отчёта.</li>
        <li>Решения по системным подрядчикам принимать на уровне договора, а не по каждому объекту отдельно.</li>
        <li>В переговорах с подрядчиком: «хотите быстрее получать деньги — подавайте акты вовремя» — это и есть рычаг ускорения, а не дополнительное финансирование.</li>
      </ul>

      <strong style="display:block;margin-top:14px">Ограничения</strong>
      <p class="note">Построено на 48 объектах капремонта школ — на другие типы объектов переносить с проверкой. Разброс вокруг эталонной кривой большой (см. график выше), поэтому «красная зона» — это ориентир для разбора, а не жёсткий автоматический триггер. Данные по подрядчикам и бюджету 2026 сверены вручную один раз и не обновляются автоматически вместе с остальным отчётом.</p>
    </div>
  </details>

  <div class="box mini">
    <strong style="display:block;margin-bottom:8px">Кого смотреть первым</strong>
    <table>
      <thead><tr><th>Школа</th><th class="r">СГ</th><th class="r">План</th><th class="r">Выпл.</th><th>Что не так</th></tr></thead>
      <tbody id="ktAttn"></tbody>
    </table>
  </div>

  <details id="secCharts">
    <summary>Графики по школе</summary>
    <div class="detail-body">
      <strong style="display:block;margin-top:14px">Готовность и выплаты</strong>
      <div class="chart"><canvas id="cScatter"></canvas></div>
      <p class="note">Точка — школа (наведите курсор, чтобы узнать название). Серые — типовой аванс (30% или 49%), не индивидуально посчитанный факт оплаты. Пунктир — общий тренд по всем точкам: он почти горизонтальный, роста почти нет.</p>
      <p class="note">Даже если убрать авансы и оставить только <span id="mF"></span> школ с индивидуальным процентом, связь всё равно не появляется: r=<span id="mG"></span>, p≈<span id="mH"></span>. А вот <strong>внутри одной школы</strong> во времени готовность и выплаты растут почти синхронно — медианная корреляция <span id="mC"></span> по <span id="mD"></span> школам (см. «Одна школа» ниже). Разница простая: «сколько уже закрыто актами по этой стройке» — не то же самое, что «у кого выше готовность по сравнению с другими школами».</p>
      <details style="margin-top:6px">
        <summary style="cursor:pointer;font-size:.84rem;color:var(--accent)">Что значат r и p?</summary>
        <ul class="brief" style="margin-top:8px">
          <li><strong>r</strong> — насколько сильно две величины растут вместе, от −1 до +1. 0 — совсем никакой связи, точки разбросаны как попало. +1 — идеальная связь: одно растёт, второе растёт ровно так же. −1 — идеальная обратная связь.</li>
          <li><strong>p</strong> — какова вероятность увидеть такое же r случайно, если на самом деле никакой связи нет вообще. Маленький p (меньше 0.05, то есть 5%) — связь, скорее всего, настоящая. Большой p — это вполне может быть просто совпадение на нашей выборке школ.</li>
        </ul>
      </details>
      <strong style="display:block;margin-top:18px">Одна школа</strong>
      <p class="note">Звёздочка (*) в списке — у школы есть хотя бы одно замечание (см. раздел «Все школы: даты и разрывы» ниже).</p>
      <div class="row">
        <select id="selSchool"></select>
        <span class="tag" id="tagR"></span>
        <span class="note" id="metaSchool" style="margin:0"></span>
      </div>
      <div class="chart"><canvas id="cTraj"></canvas></div>
      <p class="note" id="ktLine"></p>
      <p class="note">Синяя — факт, пунктир — план, зелёная — % от суммы контракта, уже выплаченной на эту дату.</p>
    </div>
  </details>

  <details id="secKt">
    <summary>Все школы: даты и разрывы</summary>
    <div class="detail-body">
      <div class="tbl-wrap">
        <table class="full">
          <thead><tr>
            <th>Школа</th><th class="r">Факт</th><th class="r">План</th><th class="r">Δ</th>
            <th class="r">Выпл.</th><th class="r">Δ</th>
            <th>Экспертиза</th><th>Старт СМР</th><th>Контракт</th><th>Замечания</th>
          </tr></thead>
          <tbody id="ktTbl"></tbody>
        </table>
      </div>
    </div>
  </details>

  <details id="secContractors">
    <summary>По подрядчикам</summary>
    <div class="detail-body">
      <p class="note" style="margin-top:14px">Сгруппировали школы по подрядчику — видно, повторяются ли проблемы у одного и того же исполнителя на разных объектах, или это разовые случаи.</p>
      <div class="tbl-wrap">
        <table class="full">
          <thead><tr>
            <th>Подрядчик</th><th class="r">Объектов</th><th class="r">Не осваивают бюджет 2026</th><th>Объекты</th>
          </tr></thead>
          <tbody id="contractorsTbl"></tbody>
        </table>
      </div>
    </div>
  </details>

  <details id="secTable">
    <summary>Краткая таблица</summary>
    <div class="detail-body">
      <div class="tbl-wrap">
        <table class="full">
          <thead><tr>
            <th>Школа</th><th class="r">СГ</th><th class="r">План</th><th class="r">Выпл.</th>
            <th class="r">Разрыв</th><th class="r">Сходятся</th>
          </tr></thead>
          <tbody id="tbl"></tbody>
        </table>
      </div>
    </div>
  </details>

  <details id="secContext">
    <summary>Если коротко</summary>
    <div class="detail-body">
      <ul class="brief" style="margin-top:14px">
        <li>Готовность и деньги считаются отдельно — высокая СГ при 30% выплат это нормально.</li>
        <li>Деньги идут после приёмки, не по цифре с мониторинга.</li>
        <li>«Стройка до экспертизы» — в КСГ экспертиза позже, а СМР уже идут.</li>
        <li>РС в этой выгрузке по школам не выделен.</li>
      </ul>
      <p class="note">Данные: файлы СГ, платежи, Simple List, КСГ+Экспертиза.</p>
    </div>
  </details>
</div>
<script>
const DATA = __DATA__;
const blue='#1a4d7a', blueL='rgba(26,77,122,.35)', green='#2d6a4f';
const charts = { scatter:false, traj:false };

function flagsHtml(arr) {
  if(!arr||!arr.length) return '—';
  return arr.map(f=>'<span class="flag'+(f.includes('стройка')||f.includes('эксперт')||f.includes('бюджет')?' warn':'')+'">'+f+'</span>').join('');
}

document.getElementById('k1').textContent = DATA.stats.n_sg_ahead + ' из ' + DATA.stats.n;
document.getElementById('k2').textContent = '+' + DATA.stats.median_gap;
document.getElementById('k4').textContent = DATA.stats.n_sg_behind_plan;
document.getElementById('k5').textContent = DATA.stats.n_smr_before_exp;

document.getElementById('mC').textContent = DATA.stats.median_r.toFixed(2);
document.getElementById('mD').textContent = DATA.stats.n_varying;
document.getElementById('mF').textContent = DATA.stats.n_no_advance;
document.getElementById('mG').textContent = DATA.stats.pearson_no_advance!=null ? DATA.stats.pearson_no_advance.toFixed(2) : '—';
document.getElementById('mH').textContent = DATA.stats.pearson_no_advance_p!=null ? DATA.stats.pearson_no_advance_p.toFixed(2) : '—';

document.getElementById('ktAttn').innerHTML = DATA.kt_attention.map(s =>
  `<tr><td title="${s.full}">${s.name}</td><td class="r">${s.sg}%</td><td class="r">${s.plan??'—'}%</td><td class="r">${s.pct??'—'}%</td><td>${flagsHtml(s.flags)}</td></tr>`
).join('');

document.getElementById('b1').textContent = DATA.stats.n_sg_behind_plan;
document.getElementById('b2').textContent = DATA.stats.n_sg_ahead;
document.getElementById('b3').textContent = DATA.stats.n_smr_before_exp;
document.getElementById('b4').textContent = DATA.stats.n_no_budget2026;
document.getElementById('k6').textContent = DATA.stats.n_no_budget2026;

document.getElementById('budgetAlert').innerHTML = DATA.budget_alert.map(s =>
  `<tr><td title="${s.full}">${s.name}</td><td class="r">${s.sg}%</td><td class="r">${s.plan2026}</td></tr>`
).join('');

document.getElementById('contractorsTbl').innerHTML = DATA.contractors.map(c => {
  const objs = c.objects.map(o => o.name + (o.no_budget2026 ? ' *' : '')).join(', ');
  const allStuck = c.objects.length > 1 && c.n_no_budget2026 === c.objects.length;
  return `<tr${allStuck ? ' style="background:#fdecea"' : ''}><td>${c.contractor}</td><td class="r">${c.objects.length}</td><td class="r">${c.n_no_budget2026 || '—'}</td><td>${objs}</td></tr>`;
}).join('');

document.getElementById('ktTbl').innerHTML = [...DATA.per].sort((a,b)=>(b.flags?.length||0)-(a.flags?.length||0)).map(p=>{
  const k = DATA.kt_dates[p.uin]||{};
  const exp = [k.exp_sl,k.exp_plan,k.exp_fact].filter(Boolean).join(' / ') || '—';
  return `<tr><td>${p.name}</td><td class="r">${p.sg}</td><td class="r">${p.plan??'—'}</td><td class="r">${p.sg_plan_gap??'—'}</td><td class="r">${p.pct??'—'}</td><td class="r">${p.sg!=null&&p.pct!=null?(p.sg-p.pct).toFixed(0):'—'}</td><td>${exp}</td><td>${k.smr_start||'—'}</td><td>${k.ctr_fact||'—'}</td><td>${flagsHtml(p.flags)}</td></tr>`;
}).join('');

document.getElementById('tbl').innerHTML = DATA.per.map(p => {
  const gap = p.pct!=null ? (p.sg-p.pct).toFixed(0) : '—';
  return `<tr><td>${p.name}</td><td class="r">${p.sg}</td><td class="r">${p.plan??'—'}</td><td class="r">${p.pct??'—'}</td><td class="r">${gap}</td><td class="r">${p.r!=null?p.r.toFixed(2):'—'}</td></tr>`;
}).join('');

const sel = document.getElementById('selSchool');
DATA.per.filter(p=>p.vary).forEach(p=>{
  const o=document.createElement('option'); o.value=p.uin;
  o.textContent=p.name + (p.flags?.length ? ' *' : ''); sel.appendChild(o);
});
sel.value = DATA.defaultUin;
let chartTraj;

function drawTraj(uin) {
  const rows = DATA.traj[uin]||[], m = DATA.per.find(p=>p.uin===uin);
  const k = DATA.kt_dates[uin]||{};
  document.getElementById('tagR').textContent = m&&m.r!=null ? 'корр. '+m.r.toFixed(2) : '';
  document.getElementById('metaSchool').textContent = m ? `${m.name}: ${m.sg}% факт, ${m.plan??'—'}% план, ${m.pct??'—'}% выпл.` : '';
  document.getElementById('ktLine').textContent = `Экспертиза ${k.exp_sl||k.exp_plan||'—'}, СМР с ${k.smr_start||'—'}, контракт ${k.ctr_fact||'—'}` + (m&&m.flags?.length ? '. '+m.flags.join(', ') : '');
  const annotations = {};
  if (k.exp_marker) {
    annotations.exp = { type:'line', xMin:k.exp_marker, xMax:k.exp_marker, borderColor:'#a04ea3', borderWidth:2, borderDash:[3,3],
      label:{ display:true, content:'Экспертиза', position:'start', backgroundColor:'#a04ea3', font:{size:10} } };
  }
  const cfg = { type:'line', data:{ labels:rows.map(r=>r.d), datasets:[
    { label:'Факт', data:rows.map(r=>r.sg), borderColor:blue, tension:.25, pointRadius:2 },
    { label:'План', data:rows.map(r=>r.plan), borderColor:blue, borderDash:[5,4], tension:.25, pointRadius:0 },
    { label:'Основные платежи', data:rows.map(r=>r.regPct), borderColor:green, tension:.25, pointRadius:2 },
    { label:'Аванс', data:rows.map(r=>r.advPct), borderColor:green, borderDash:[2,2], tension:.25, pointRadius:0 }
  ]}, options:{ responsive:true, maintainAspectRatio:false, plugins:{legend:{position:'bottom'}, datalabels:{display:false}, annotation:{annotations}}, scales:{ y:{min:0,max:100} } } };
  if(chartTraj) chartTraj.destroy(); chartTraj = new Chart(document.getElementById('cTraj'), cfg);
}
sel.onchange = e => drawTraj(e.target.value);

function initDetailCharts() {
  if(charts.scatter) return;
  charts.scatter = true;
  const pts = DATA.cross.filter(s=>s.pct!=null);
  const mk2 = s => ({x:s.pct,y:s.sg,name:s.name,adv:s.advance});
  const line=[]; for(let x=25;x<=95;x+=5) line.push({x,y:DATA.reg.a+DATA.reg.b*x});
  new Chart(document.getElementById('cScatter'), {
    type:'scatter',
    data:{ datasets:[
      { label:'Свой процент', data:pts.filter(s=>!s.advance).map(mk2), backgroundColor:'rgba(26,77,122,.75)', pointRadius:4 },
      { label:'Аванс (30% или 49%)', data:pts.filter(s=>s.advance).map(mk2), backgroundColor:'rgba(140,140,140,.6)', pointRadius:4 },
      { label:'Тренд по всем школам', data:line, type:'line', borderColor:blue, borderDash:[5,4], pointRadius:0 }
    ]},
    options:{ responsive:true, maintainAspectRatio:false, plugins:{legend:{position:'bottom'}, datalabels:{display:false},
      tooltip:{callbacks:{label:c=>c.raw.name ? `${c.raw.name}: выплаты ${c.raw.x}%${c.raw.adv?' (аванс)':''}, СГ ${c.raw.y}%` : 'тренд'}}},
      scales:{ x:{title:{display:true,text:'Выплаты, %'},min:20,max:100}, y:{title:{display:true,text:'СГ, %'},min:70,max:100} } }
  });
  drawTraj(sel.value);
  charts.traj = true;
}

function initStage() {
  const s = DATA.stage_analysis;
  new Chart(document.getElementById('cStage'), {
    type:'line',
    data:{ labels:s.map(x=>x.stage+'%'), datasets:[
      { label:'СГ обгоняет выплаты, п.п.', data:s.map(x=>x.pay_gap), borderColor:green, backgroundColor:green, tension:.2, pointRadius:4 },
      { label:'Факт отстаёт от плана, п.п.', data:s.map(x=>x.plan_gap), borderColor:blue, backgroundColor:blue, borderDash:[5,4], tension:.2, pointRadius:4 }
    ]},
    options:{ responsive:true, maintainAspectRatio:false, plugins:{legend:{position:'bottom'}, datalabels:{display:false},
      tooltip:{callbacks:{afterLabel:c=>'школ на этом этапе: '+s[c.dataIndex].n}}},
      scales:{ x:{title:{display:true,text:'Этап готовности (СГ)'}}, y:{title:{display:true,text:'Разрыв, п.п.'}} } }
  });
  const first = s.find(x=>x.n>=10), last = [...s].reverse().find(x=>x.n>=10);
  document.getElementById('stageNote').textContent = first && last
    ? `На старте (${first.stage}% готовности) деньги обгоняют стройку на ${Math.abs(first.pay_gap)} п.п., а к ${last.stage}% готовности стройка уже обгоняет деньги на ${last.pay_gap} п.п. Разрыв не постоянный — он растёт по ходу стройки, особенно заметно после 60–70% готовности. Пунктирная линия (план) показывает обратное: сильнее всего от своего же плана школы отстают в середине стройки, а к концу почти нагоняют.`
    : '';
}

function initMethod() {
  const s = DATA.stage_analysis;
  const late = [...s].reverse().find(x=>x.n>=10 && x.stage>=80) || [...s].reverse().find(x=>x.n>=10);
  document.getElementById('mth90').textContent = late ? late.pay_gap : '—';
  document.getElementById('mth0').textContent = DATA.stats.n_no_budget2026;

  const stuck = DATA.contractors.filter(c => c.objects.length > 1 && c.n_no_budget2026 === c.objects.length);
  document.getElementById('mthContr').textContent = stuck.length
    ? `Пример: у ${stuck.length>1?'подрядчиков':'подрядчика'} ${stuck.map(c=>c.contractor+' ('+c.n_no_budget2026+' из '+c.objects.length+' объектов)').join(', ')} бюджет 2026 года не осваивает каждый без исключения объект — это уже не единичный случай, а системный признак: если у подрядчика все объекты в одном состоянии, разбираться нужно с самим подрядчиком, а не перебирать его объекты по одному.`
    : 'На этой выгрузке подрядчиков, у которых стоят все объекты сразу, не найдено — но при появлении такого паттерна это приоритетный сигнал.';
}

initStage();
initMethod();
document.getElementById('secCharts').addEventListener('toggle', e => { if(e.target.open) initDetailCharts(); });
</script>
</body>
</html>
"""


def main():
    payload = load_data()
    html = HTML_TEMPLATE.replace("__DATA__", json.dumps(payload, ensure_ascii=False))
    out = ROOT / "index.html"
    out.write_text(html, encoding="utf-8")
    (ROOT / "sg-pay-analysis.html").write_text(html, encoding="utf-8")
    print(f"OK: {out} ({out.stat().st_size} bytes, {payload['stats']['n']} schools)")


if __name__ == "__main__":
    main()
